"""异常点查询控制器"""
import csv
import io
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.config.database import get_db
from app.config.response import success, BusinessException
from app.models.anomaly import AnomalyRecord
from app.models.task import Task
from app.services.data_service import DataService
from app.config.logging import get_logger

logger = get_logger("routers.anomaly")
anomaly_router = APIRouter(tags=["异常管理"])

INDICATOR_SHORT = ["chl", "odo", "temp", "ph", "turb"]


@anomaly_router.get("/api/anomalies", summary="全部异常点列表（跨任务）")
async def get_all_anomalies(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=1000),
    indicator: str = Query(None),
    task_id: str = Query(None),
    method: str = Query(None, description="检测方法: threshold / isolation_forest"),
    db: AsyncSession = Depends(get_db),
):
    base = select(AnomalyRecord)
    if indicator and indicator in INDICATOR_SHORT:
        base = base.where(AnomalyRecord.indicator == indicator)
    if task_id:
        base = base.where(AnomalyRecord.task_id == task_id)
    if method and method in ("threshold", "isolation_forest"):
        base = base.where(AnomalyRecord.method == method)

    count_q = select(func.count()).select_from(base.subquery())
    result = await db.execute(count_q)
    total = result.scalar()

    result = await db.execute(
        base.order_by(AnomalyRecord.id.desc()).offset((page - 1) * page_size).limit(page_size)
    )
    data = [
        {"id": r.id, "task_id": r.task_id, "lon": r.lon, "lat": r.lat, "depth": r.depth,
         "indicator": r.indicator, "value": r.value, "method": r.method}
        for r in result.scalars().all()
    ]
    return success(datas={"total": total, "items": data})


@anomaly_router.get("/api/task/{task_id}/anomalies", summary="异常点列表")
async def get_anomalies(
    task_id: str,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=1000),
    db: AsyncSession = Depends(get_db),
):
    task = await DataService.get_task(db, task_id)
    if not task:
        raise BusinessException(msg="任务不存在", code=404)

    result = await db.execute(
        select(func.count()).select_from(AnomalyRecord).where(AnomalyRecord.task_id == task_id)
    )
    total = result.scalar()

    result = await db.execute(
        select(AnomalyRecord).where(AnomalyRecord.task_id == task_id)
        .offset((page - 1) * page_size).limit(page_size)
    )
    data = [
        {"id": r.id, "lon": r.lon, "lat": r.lat, "depth": r.depth,
         "indicator": r.indicator, "value": r.value, "method": r.method}
        for r in result.scalars().all()
    ]
    return success(datas={"total": total, "items": data})


INDICATOR_CN = {"chl": "叶绿素", "odo": "溶解氧", "temp": "水温", "ph": "pH", "turb": "浊度"}
METHOD_CN = {"threshold": "统计阈值", "isolation_forest": "孤立森林"}


@anomaly_router.get("/api/task/{task_id}/anomalies/export", summary="导出异常数据")
async def export_anomalies(
    task_id: str,
    format: str = Query("csv", description="导出格式: csv 或 xlsx"),
    db: AsyncSession = Depends(get_db),
):
    task = await DataService.get_task(db, task_id)
    if not task:
        raise BusinessException(msg="任务不存在", code=404)

    result = await db.execute(
        select(AnomalyRecord).where(AnomalyRecord.task_id == task_id)
    )
    rows = result.scalars().all()

    if format == "xlsx":
        return _export_xlsx(rows, task_id)
    return _export_csv(rows, task_id)


def _export_csv(rows, task_id):
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=["lon", "lat", "depth", "indicator", "value", "method"])
    w.writeheader()
    w.writerows([
        {"lon": r.lon, "lat": r.lat, "depth": r.depth,
         "indicator": INDICATOR_CN.get(r.indicator, r.indicator),
         "value": r.value, "method": METHOD_CN.get(r.method, r.method)}
        for r in rows
    ])
    return StreamingResponse(
        io.BytesIO(buf.getvalue().encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={task_id}_anomalies.csv"},
    )


def _export_xlsx(rows, task_id):
    wb = Workbook()
    ws = wb.active
    ws.title = "异常数据"

    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="409EFF", end_color="409EFF", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    cell_font = Font(name="微软雅黑", size=10)
    cell_align = Alignment(horizontal="center", vertical="center")
    danger_font = Font(name="微软雅黑", size=10, color="F56C6C", bold=True)
    danger_fill = PatternFill(start_color="FEF0F0", end_color="FEF0F0", fill_type="solid")

    headers = ["经度", "纬度", "深度(m)", "指标", "异常值", "检测方法", "阈值下限", "阈值上限"]
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for i, r in enumerate(rows, 2):
        values = [
            f"{r.lon:.5f}", f"{r.lat:.5f}", f"{r.depth:.1f}",
            INDICATOR_CN.get(r.indicator, r.indicator),
            f"{r.value:.4f}",
            METHOD_CN.get(r.method, r.method),
            f"{r.threshold_low:.2f}" if r.threshold_low is not None else "-",
            f"{r.threshold_high:.2f}" if r.threshold_high is not None else "-",
        ]
        is_anomaly = r.method in ("threshold", "isolation_forest")
        for j, v in enumerate(values, 1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.font = danger_font if is_anomaly else cell_font
            cell.fill = danger_fill if is_anomaly else PatternFill()
            cell.alignment = cell_align
            cell.border = thin_border

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={task_id}_anomalies.xlsx"},
    )
